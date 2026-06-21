"""Ekspor laporan proyek ke Excel (.xlsx) via openpyxl.

Dua sheet: "Logbook" dan "Keuangan". Header & total diberi gaya sederhana.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.laporan.data import LaporanData

_HEADER_FILL = PatternFill("solid", fgColor="334155")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _tulis_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def buat_excel_laporan(data: LaporanData) -> bytes:
    wb = Workbook()

    # ── Sheet identitas + Logbook ──
    ws = wb.active
    ws.title = "Logbook"
    ws.append(["Judul Proyek", data.proyek.judul])
    ws.append(["Ketua Peneliti", data.ketua_nama])
    ws.append(["Status", data.proyek.status.value])
    ws.append([])
    _tulis_header(ws, ["Tanggal", "Tugas", "Pelaksana", "Kegiatan", "Durasi"])
    for b in data.logbook:
        ws.append([b.tanggal, b.tugas, b.pelaksana, b.kegiatan, b.durasi])
    for col, lebar in zip("ABCDE", (18, 25, 20, 45, 12)):
        ws.column_dimensions[col].width = lebar

    # ── Sheet Keuangan ──
    wk = wb.create_sheet("Keuangan")
    _tulis_header(wk, ["Kategori", "Dialokasikan", "Terpakai", "Sisa", "Sisa %"])
    rk = data.ringkasan_keuangan
    for k in rk.kategori:
        wk.append([
            k.kategori, float(k.dialokasikan), float(k.terpakai),
            float(k.sisa), k.persen_sisa,
        ])
    total_row = ["TOTAL", float(rk.total_dialokasikan), float(rk.total_terpakai),
                 float(rk.total_sisa), ""]
    wk.append(total_row)
    for cell in wk[wk.max_row]:
        cell.font = Font(bold=True)
    # Format rupiah untuk kolom nominal.
    for row in wk.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = '#,##0'
    for col, lebar in zip("ABCDE", (28, 16, 16, 16, 10)):
        wk.column_dimensions[col].width = lebar

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
